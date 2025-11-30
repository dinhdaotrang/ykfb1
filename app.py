import streamlit as st
import pandas as pd
import os
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io

# Configuration
CSV_FILE = "feedback_data.csv"

def load_feedback_data():
    """Load existing feedback data from CSV file."""
    if os.path.exists(CSV_FILE):
        try:
            df = pd.read_csv(CSV_FILE)
            return df
        except Exception as e:
            st.error(f"Lỗi khi tải dữ liệu: {e}")
            return pd.DataFrame(columns=["timestamp", "name", "rating", "feedback"])
    else:
        return pd.DataFrame(columns=["timestamp", "name", "rating", "feedback"])

def save_feedback(name, rating, feedback):
    """Save feedback to CSV file."""
    # Create new entry
    new_entry = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "name": name if name else "Ẩn danh",
        "rating": rating,
        "feedback": feedback
    }
    
    # Load existing data
    df = load_feedback_data()
    
    # Append new entry
    new_df = pd.DataFrame([new_entry])
    df = pd.concat([df, new_df], ignore_index=True)
    
    # Save to CSV
    df.to_csv(CSV_FILE, index=False)
    
    return True

def export_to_excel_tcvn(df):
    """Export feedback data to Excel file with TCVN format (Times New Roman, size 13)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ý Kiến"
    
    # TCVN Font and Style Settings
    tcvn_font = Font(name='Times New Roman', size=13)
    header_font = Font(name='Times New Roman', size=13, bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header fill color (light gray)
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    # Tiêu đề
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "BÁO CÁO Ý KIẾN NGƯỜI DÙNG"
    title_cell.font = Font(name='Times New Roman', size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Thông tin báo cáo
    ws['A2'] = f"Ngày xuất báo cáo: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws['A2'].font = tcvn_font
    ws.merge_cells('A2:E2')
    ws['A2'].alignment = left_alignment
    
    # Khoảng trống
    ws.row_dimensions[3].height = 10
    
    # Headers
    headers = ["STT", "Ngày & Giờ", "Tên", "Đánh Giá", "Ý Kiến"]
    header_row = 4
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 8   # STT
    ws.column_dimensions['B'].width = 20  # Ngày & Giờ
    ws.column_dimensions['C'].width = 25  # Tên
    ws.column_dimensions['D'].width = 12  # Đánh Giá
    ws.column_dimensions['E'].width = 50  # Ý Kiến
    
    # Data rows
    df_sorted = df.copy()
    df_sorted = df_sorted.sort_values('timestamp').reset_index(drop=True)
    
    for idx, row in df_sorted.iterrows():
        data_row = header_row + idx + 1
        
        # STT
        ws.cell(row=data_row, column=1, value=idx + 1)
        ws.cell(row=data_row, column=1).font = tcvn_font
        ws.cell(row=data_row, column=1).alignment = center_alignment
        ws.cell(row=data_row, column=1).border = thin_border
        
        # Ngày & Giờ
        timestamp = row['timestamp']
        ws.cell(row=data_row, column=2, value=timestamp)
        ws.cell(row=data_row, column=2).font = tcvn_font
        ws.cell(row=data_row, column=2).alignment = center_alignment
        ws.cell(row=data_row, column=2).border = thin_border
        
        # Tên
        ws.cell(row=data_row, column=3, value=row['name'])
        ws.cell(row=data_row, column=3).font = tcvn_font
        ws.cell(row=data_row, column=3).alignment = left_alignment
        ws.cell(row=data_row, column=3).border = thin_border
        
        # Đánh Giá
        rating_value = f"{row['rating']} ⭐"
        ws.cell(row=data_row, column=4, value=rating_value)
        ws.cell(row=data_row, column=4).font = tcvn_font
        ws.cell(row=data_row, column=4).alignment = center_alignment
        ws.cell(row=data_row, column=4).border = thin_border
        
        # Ý Kiến
        ws.cell(row=data_row, column=5, value=row['feedback'])
        ws.cell(row=data_row, column=5).font = tcvn_font
        ws.cell(row=data_row, column=5).alignment = left_alignment
        ws.cell(row=data_row, column=5).border = thin_border
        
        # Set row height for better readability
        ws.row_dimensions[data_row].height = 25
    
    # Set header row height
    ws.row_dimensions[header_row].height = 30
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()

def main():
    """Main Streamlit application."""
    st.set_page_config(
        page_title="Thu Thập Ý Kiến",
        page_icon="💬",
        layout="centered"
    )
    
    # Title and description
    st.title("💬 Thu Thập Ý Kiến")
    st.markdown("---")
    st.markdown("Chúng tôi rất trân trọng ý kiến của bạn! Vui lòng chia sẻ phản hồi của bạn bên dưới.")
    
    # Feedback form
    with st.form("feedback_form", clear_on_submit=True):
        st.subheader("Gửi Ý Kiến Của Bạn")
        
        # Name field (optional)
        name = st.text_input(
            "Tên (Tùy chọn)",
            placeholder="Nhập tên của bạn hoặc để trống nếu muốn ẩn danh"
        )
        
        # Rating field (required)
        rating = st.selectbox(
            "Đánh giá *",
            options=["", "1 ⭐", "2 ⭐⭐", "3 ⭐⭐⭐", "4 ⭐⭐⭐⭐", "5 ⭐⭐⭐⭐⭐"],
            help="Vui lòng chọn mức đánh giá từ 1 đến 5"
        )
        
        # Feedback text area
        feedback = st.text_area(
            "Ý Kiến Của Bạn",
            placeholder="Chia sẻ suy nghĩ, đề xuất hoặc nhận xét của bạn tại đây...",
            height=150
        )
        
        # Submit button
        submitted = st.form_submit_button("Gửi Ý Kiến", type="primary")
        
        if submitted:
            # Validation
            if not rating:
                st.error("⚠️ Vui lòng chọn mức đánh giá trước khi gửi.")
            elif not feedback.strip():
                st.warning("⚠️ Vui lòng nhập ý kiến của bạn.")
            else:
                # Extract numeric rating from selection
                rating_value = rating.split()[0] if rating else None
                
                # Save feedback
                try:
                    save_feedback(name, rating_value, feedback)
                    st.success("✅ Cảm ơn bạn! Ý kiến của bạn đã được gửi thành công.")
                except Exception as e:
                    st.error(f"❌ Đã xảy ra lỗi khi lưu ý kiến của bạn: {e}")
    
    st.markdown("---")
    
    # Display collected feedback
    st.subheader("📊 Ý Kiến Đã Thu Thập")
    
    df = load_feedback_data()
    
    if df.empty:
        st.info("Chưa có ý kiến nào được thu thập. Hãy là người đầu tiên chia sẻ ý kiến của bạn!")
    else:
        # Display statistics
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Tổng Số Phản Hồi", len(df))
        with col2:
            avg_rating = df["rating"].astype(float).mean()
            st.metric("Đánh Giá Trung Bình", f"{avg_rating:.1f} ⭐")
        with col3:
            st.metric("Ý Kiến Mới Nhất", df.iloc[-1]["timestamp"].split()[0] if not df.empty else "N/A")
        
        st.markdown("")
        
        # Display feedback table
        # Create a more readable display
        display_df = df.copy()
        display_df = display_df.rename(columns={
            "timestamp": "Ngày & Giờ",
            "name": "Tên",
            "rating": "Đánh Giá",
            "feedback": "Ý Kiến"
        })
        
        # Format rating to show stars
        display_df["Đánh Giá"] = display_df["Đánh Giá"].astype(str) + " ⭐"
        
        # Reverse order to show latest first
        display_df = display_df.iloc[::-1].reset_index(drop=True)
        
        # Display table
        st.dataframe(
            display_df,
            use_container_width=True,
            hide_index=True
        )
        
        # Download button for Excel (TCVN format)
        try:
            excel_data = export_to_excel_tcvn(df)
            st.download_button(
                label="📥 Tải Xuống Dữ Liệu Ý Kiến (Excel)",
                data=excel_data,
                file_name=f"y_kien_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"❌ Lỗi khi tạo file Excel: {e}")

if __name__ == "__main__":
    main()

